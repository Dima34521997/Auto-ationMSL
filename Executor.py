import openpyxl
import time
from datetime import date

from openpyxl.styles import (Alignment,
                             Border,
                             Side)


def execute(ds, table):
    begin = time.time()
    wb = openpyxl.load_workbook(table)
    ws = wb.active
    end = time.time()

    print("Время открытия таблицы:", end - begin)

    # ws = wb.get_sheet_by_name('ПИ 026-04')
    free_cell = 2

    while ws[f'A{free_cell}'].value is not None:
        free_cell += 1

    try:
        ds.msl_number = int(ws[f'A{free_cell - 1}'].value) + 1

    except ValueError:
        pass

    if ds.device_name == 'МСЛ ББ':
        """
        Для батарейных блоков число позиций на МСЛ и общее кол-во
        позиций х4, т. к. в блоке 4 аккумулятора
        """
        ds.per_one_msl *= 4
        ds.amount *= 4

    integer_msl: int = ds.amount // ds.per_one_msl
    """Число целых МСЛок"""

    residue: int = ds.amount % ds.per_one_msl
    """Остаточные позиции в последнюю МСЛ"""

    # borders = Border(left=Side(style='thin'),
    #                  right=Side(style='thin'),
    #                  top=Side(style='thin'),
    #                  bottom=Side(style='thin'))

    if residue == 0:
        """Если все МСЛ полные"""
        for i in range(integer_msl):
            ws[f'A{free_cell + i}'] = ds.msl_number + i
            ws[f'A{free_cell + i}'].alignment = Alignment(horizontal='center')

            # ws[f'B{free_cell + i}'] = ds.device_name
            # ws[f'B{free_cell + i}'].alignment = Alignment(horizontal='center')

            ws[f'B{free_cell + i}'] = ds.per_one_msl if ds.device_name != 'МСЛ ББ' else ds.per_one_msl // 4
            ws[f'B{free_cell + i}'].alignment = Alignment(horizontal='center')

            if ds.first - (ds.first + ds.per_one_msl - 1) != 0:
                """Если число позиций в МСЛ > 0"""
                ws[f'C{free_cell + i}'] = (f'{ds.first}  '
                                           f'- {ds.first + ds.per_one_msl - 1}')
            else:
                """Если число позиций в МСЛ = 1"""
                ws[f'C{free_cell + i}'] = f'{ds.first}'

            ws[f'D{free_cell + i}'] = date.today().strftime("%d.%m.%Y")
            ws[f'D{free_cell + i}'].alignment = Alignment(horizontal='center')

            ws[f'E{free_cell + i}'] = ds.master_name
            ws[f'E{free_cell + i}'].alignment = Alignment(horizontal='center')

            ws[f'F{free_cell + i}'] = ds.contract
            ws[f'F{free_cell + i}'].alignment = Alignment(horizontal='center')

            ds.first += ds.per_one_msl

    else:
        """Если будут неполные МСЛ (число позиций в МСЛ < per_one_msl)"""
        count = 0
        for count in range(integer_msl + 1):
            ws[f'A{free_cell + count}'] = ds.msl_number + count
            ws[f'A{free_cell + count}'].alignment = Alignment(horizontal='center')

            # ws[f'B{free_cell + count}'] = ds.device_name
            # ws[f'B{free_cell + count}'].alignment = Alignment(horizontal='center')

            ws[f'B{free_cell + count}'] = ds.per_one_msl if ds.device_name != 'МСЛ ББ' else ds.per_one_msl // 4
            ws[f'B{free_cell + count}'].alignment = Alignment(horizontal='center')

            ws[f'C{free_cell + count}'] = (f'{ds.first}'
                                           f' - {ds.first + ds.per_one_msl - 1}')
            ws[f'C{free_cell + count}'].alignment = Alignment(horizontal='center')

            ws[f'D{free_cell + count}'] = date.today().strftime("%d.%m.%Y")
            ws[f'D{free_cell + count}'].alignment = Alignment(horizontal='center')

            ws[f'E{free_cell + count}'] = ds.master_name
            ws[f'E{free_cell + count}'].alignment = Alignment(horizontal='center')

            ws[f'F{free_cell + count}'] = ds.contract
            ws[f'F{free_cell + count}'].alignment = Alignment(horizontal='center')

            ds.first += ds.per_one_msl

        """Добавляем остаток отдельно"""
        ws[f'A{free_cell + count}'] = ds.msl_number + count
        ws[f'A{free_cell + count}'].alignment = Alignment(horizontal='center')

        # ws[f'B{free_cell + count}'] = ds.device_name
        # ws[f'B{free_cell + count}'].alignment = Alignment(horizontal='center')

        ws[f'B{free_cell + count}'] = residue if ds.device_name != 'МСЛ ББ' else residue // 4
        ws[f'B{free_cell + count}'].alignment = Alignment(horizontal='center')

        kurwa = ds.first - ds.per_one_msl - ds.first + residue + ds.per_one_msl - 1
        if kurwa != 0:
            """Если число позиций в МСЛ > 0"""
            ws[f'C{free_cell + count}'] = (f'{ds.first - ds.per_one_msl}'
                                           f' - {ds.first + residue - ds.per_one_msl - 1}')
        else:
            """Если число позиций в МСЛ = 1"""
            ws[f'C{free_cell + count}'] = f'{ds.first - ds.per_one_msl}'

        ws[f'C{free_cell + count}'].alignment = Alignment(horizontal='center')

        ws[f'D{free_cell + count}'] = date.today().strftime("%d.%m.%Y")
        ws[f'D{free_cell + count}'].alignment = Alignment(horizontal='center')

        ws[f'E{free_cell + count}'] = ds.master_name
        ws[f'E{free_cell + count}'].alignment = Alignment(horizontal='center')

        ws[f'F{free_cell + count}'] = ds.contract
        ws[f'F{free_cell + count}'].alignment = Alignment(horizontal='center')
    # ds.refresh_data()

    begin = time.time()

    wb.save(table)

    end = time.time()

    print("Время закрытия таблицы:", end - begin)

    # wb.save(table)
    return True
